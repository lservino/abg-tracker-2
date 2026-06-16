"""
ABG Notice Uploader — Streamlit App
Visually matches the HTML tool.
"""

import base64, io, json, re
from datetime import date, datetime, timedelta
from copy import copy

import anthropic
import openpyxl
import pandas as pd
import streamlit as st

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ABG Notice Tracker — Upload Tool",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── CSS — mirrors the HTML tool exactly ──────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Fira+Mono&display=swap');

:root {
  --navy:        #0f1e36;
  --navy-mid:    #1a3055;
  --slate:       #2e4a6e;
  --steel:       #4a6fa0;
  --sky:         #6b9fd4;
  --ice:         #e8f0f9;
  --white:       #ffffff;
  --green:       #1a7a4a;
  --green-light: #e6f4ec;
  --amber:       #c47a00;
  --red:         #b83232;
  --red-light:   #fdf0f0;
  --border:      #cbd8e8;
  --text:        #1a2a3a;
  --muted:       #6b7f96;
}

/* Hide Streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 0 !important; max-width: 100% !important; }
section[data-testid="stSidebar"] { display: none; }

/* Full-width header */
.abg-header {
  background: var(--navy);
  color: white;
  padding: 20px 32px;
  display: flex;
  align-items: center;
  gap: 16px;
  border-bottom: 3px solid var(--sky);
  margin-bottom: 0;
}
.abg-logo {
  width: 40px; height: 40px;
  background: var(--sky);
  border-radius: 6px;
  display: flex; align-items: center; justify-content: center;
  font-weight: 800; font-size: 18px; color: var(--navy);
  letter-spacing: -1px; flex-shrink: 0;
}
.abg-header h1 { font-size: 18px; font-weight: 700; letter-spacing: .02em; margin: 0; }
.abg-header p  { font-size: 12px; color: var(--sky); margin: 2px 0 0 0; }
.abg-badge {
  margin-left: auto;
  background: var(--navy-mid);
  border: 1px solid var(--slate);
  color: var(--sky);
  font-size: 11px;
  padding: 4px 10px;
  border-radius: 4px;
  font-family: 'Fira Mono', monospace;
}

/* Main content area */
.abg-main { max-width: 960px; margin: 0 auto; padding: 32px 24px; }

/* Step cards */
.step-card {
  background: var(--white);
  border: 1px solid var(--border);
  border-radius: 10px;
  padding: 20px;
  height: 100%;
}
.step-title {
  font-size: 13px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .08em;
  color: var(--slate);
  margin-bottom: 10px;
}

/* Queue items */
.queue-item {
  background: var(--white);
  border: 1px solid var(--border);
  border-radius: 6px;
  padding: 9px 12px;
  display: flex; align-items: center; gap: 10px;
  font-size: 13px; margin-bottom: 6px;
}
.queue-fname { font-weight: 600; flex: 1; color: var(--text); }
.queue-badge {
  font-size: 10px; background: var(--ice);
  color: var(--slate); padding: 2px 7px;
  border-radius: 3px; white-space: nowrap;
}

/* Tracker zone */
.tracker-loaded {
  border: 2px solid var(--green);
  background: var(--green-light);
  color: var(--green);
  border-radius: 8px;
  padding: 14px 16px;
  font-size: 13px;
  font-weight: 600;
  text-align: center;
  margin-bottom: 8px;
}

/* Progress / log */
.log-box {
  background: var(--navy);
  border-radius: 8px;
  padding: 14px 16px;
  font-family: 'Fira Mono', monospace;
  font-size: 11px;
  color: #8bb8d8;
  max-height: 140px;
  overflow-y: auto;
  margin-top: 10px;
  line-height: 1.6;
}
.log-ok   { color: #5fca8a; }
.log-warn { color: #f5c76e; }
.log-err  { color: #f07070; }

/* Summary strip */
.summary-strip { display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }
.summary-card {
  background: var(--white);
  border: 1px solid var(--border);
  border-radius: 8px;
  padding: 12px 18px;
  flex: 1; min-width: 120px; text-align: center;
}
.summary-label {
  font-size: 11px; color: var(--muted);
  text-transform: uppercase; letter-spacing: .06em;
  margin-bottom: 4px;
}
.summary-value {
  font-size: 24px; font-weight: 800;
  color: var(--navy);
  font-family: 'Fira Mono', monospace;
}

/* Results table header */
.results-header {
  background: var(--navy);
  color: white;
  padding: 14px 20px;
  border-radius: 10px 10px 0 0;
  display: flex; align-items: center; justify-content: space-between;
}
.results-header h2 { font-size: 14px; font-weight: 700; letter-spacing: .04em; margin: 0; }
.results-header span { font-size: 12px; color: var(--sky); }

/* Download section */
.download-section {
  background: var(--green-light);
  border: 1px solid #b3dfc5;
  border-radius: 10px;
  padding: 20px 24px;
  margin-bottom: 24px;
}
.download-section h3 { font-size: 15px; font-weight: 700; color: var(--green); margin-bottom: 4px; }
.download-section p  { font-size: 13px; color: var(--text); margin: 0; }

/* Tracker status panel */
.ts-header {
  background: var(--navy);
  color: white;
  padding: 12px 20px;
  border-radius: 10px 10px 0 0;
  font-size: 14px; font-weight: 700; letter-spacing: .04em;
}

.tip { font-size: 11px; color: var(--muted); margin-top: 4px; }

/* Override Streamlit button colours */
div[data-testid="stButton"] > button[kind="primary"] {
  background: var(--navy) !important;
  border: none !important;
  font-weight: 700 !important;
  letter-spacing: .04em !important;
}
div[data-testid="stButton"] > button[kind="primary"]:hover {
  background: var(--slate) !important;
}
div[data-testid="stButton"] > button[kind="secondary"] {
  background: var(--white) !important;
  color: var(--slate) !important;
  border: 1px solid var(--border) !important;
}
div[data-testid="stDownloadButton"] > button {
  background: var(--green) !important;
  color: white !important;
  border: none !important;
  font-weight: 700 !important;
}
div[data-testid="stDownloadButton"] > button:hover {
  background: #155e38 !important;
}

/* Expander */
div[data-testid="stExpander"] summary {
  background: var(--navy);
  color: white;
  border-radius: 8px;
  font-weight: 700;
  font-size: 14px;
  letter-spacing: .04em;
}

/* File uploader */
div[data-testid="stFileUploader"] {
  border: 2px dashed var(--border);
  border-radius: 8px;
  background: var(--ice);
}

/* Metrics */
div[data-testid="stMetric"] {
  background: var(--white);
  border: 1px solid var(--border);
  border-radius: 8px;
  padding: 12px 18px;
}
div[data-testid="stMetricLabel"] { color: var(--muted) !important; font-size: 11px !important; text-transform: uppercase; letter-spacing: .06em; }
div[data-testid="stMetricValue"] { color: var(--navy) !important; font-family: 'Fira Mono', monospace; }

/* Page background */
.stApp { background: var(--ice) !important; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
TRACKER_B64_KEY  = "abg_tracker_b64"
TRACKER_DATE_KEY = "abg_tracker_date"
MODEL            = "claude-sonnet-4-6"
COL_HEADERS = [
    "File #","Completed","Date Sent","Entity Name","Jurisdiction",
    "Tax Year","Firm","Type of Notice","Notice","Status",
    "$ Issue","Adjustment","Amount Due","Refund","Notes",
    "ABG Action Items","EY Action Items",
]

# ── Session state ─────────────────────────────────────────────────────────────
for k, v in {
    "pdf_files": [], "extracted_rows": [],
    "tracker_wb": None, "tracker_label": "",
    "next_file_num": 578, "logs": [],
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Helpers ───────────────────────────────────────────────────────────────────
def log(msg, level="info"):
    st.session_state.logs.append((level, msg))

def load_wb(raw: bytes):
    return openpyxl.load_workbook(io.BytesIO(raw))

def get_next_num(ws) -> int:
    last = 575
    for row in ws.iter_rows(min_row=2, values_only=True):
        fnum = row[0]
        if fnum and any(v for v in row[1:17] if v is not None):
            if isinstance(fnum, (int, float)) and int(fnum) > last:
                last = int(fnum)
    return last + 1

def find_row(ws, filenum):
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == filenum:
            return i
    return None

def find_style_row(ws):
    best = 2
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[2] is not None and row[3] is not None:
            best = i
    return best

def parse_date(s):
    if not s: return None
    s = str(s).strip()
    for fmt in ("%m-%d-%y","%m/%d/%y","%m-%d-%Y","%m/%d/%Y","%Y-%m-%d","%B %d, %Y","%b %d, %Y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: pass
    return None

def wb_bytes(wb):
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def load_tracker(raw: bytes, label: str):
    wb = load_wb(raw)
    ws = wb["Notices"]
    st.session_state.tracker_wb    = wb
    st.session_state.tracker_label = label
    st.session_state.next_file_num = get_next_num(ws)
    st.session_state[TRACKER_B64_KEY]  = base64.b64encode(raw).decode()
    st.session_state[TRACKER_DATE_KEY] = date.today().strftime("%m-%d-%y")
    for i, f in enumerate(st.session_state.pdf_files):
        f["num"] = st.session_state.next_file_num + i
    log(f"Tracker loaded — next file # is {st.session_state.next_file_num}", "ok")

# ── Restore tracker from session if wb not loaded ─────────────────────────────
if st.session_state.tracker_wb is None and st.session_state.get(TRACKER_B64_KEY):
    wb = load_wb(base64.b64decode(st.session_state[TRACKER_B64_KEY]))
    st.session_state.tracker_wb    = wb
    st.session_state.next_file_num = get_next_num(wb["Notices"])
    st.session_state.tracker_label = f"Saved ({st.session_state.get(TRACKER_DATE_KEY,'?')})"

# ── Extraction prompt ─────────────────────────────────────────────────────────
PROMPT = """You are a tax notice extraction assistant for Authentic Brands Group (ABG).
Extract fields from this tax notice and return ONLY valid JSON, no markdown, no preamble:
{
  "entityName": "Legal name of the taxpayer",
  "jurisdiction": "Federal, or 2-letter state code e.g. FL NY",
  "taxPeriods": "Tax year(s) comma-separated e.g. 2022, 2023",
  "noticeType": "One of: Informational, Assessment, Audit, Penalty, Request for Information, Submit Information",
  "noticeDesc": "One sentence summary",
  "noticeDate": "Date as MM-DD-YY",
  "issueAmt": "Dollar amount at issue or empty string",
  "amountDue": "Amount due or empty string",
  "abgAction": "What ABG needs to do or empty string",
  "notes": "Other relevant details or empty string"
}"""

def extract(pdf_bytes: bytes) -> dict:
    client = anthropic.Anthropic()
    b64 = base64.b64encode(pdf_bytes).decode()
    msg = client.messages.create(
        model=MODEL, max_tokens=1000,
        messages=[{"role":"user","content":[
            {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}},
            {"type":"text","text":PROMPT}
        ]}]
    )
    raw = re.sub(r"^```[a-z]*\n?","",msg.content[0].text.strip())
    raw = re.sub(r"\n?```$","",raw)
    return json.loads(raw)

def write_rows(wb, rows):
    ws = wb["Notices"]
    ref = find_style_row(ws)
    def get_ref(c): return ws.cell(row=ref, column=c+1)
    for row in rows:
        r = find_row(ws, row["fileNum"]) or (ws.max_row + 1)
        def w(c, val, fmt=None):
            if val is None or val == "": return
            cell = ws.cell(row=r, column=c+1, value=val)
            rc = get_ref(c)
            if rc.number_format and rc.number_format != "General":
                cell.number_format = rc.number_format
            if fmt: cell.number_format = fmt
            if rc.font: cell.font = copy(rc.font)
        w(0, row["fileNum"])
        d = parse_date(row.get("noticeDate",""))
        if d: w(2, d, "mm-dd-yy")
        w(3, row.get("entityName",""))
        w(4, row.get("jurisdiction",""))
        w(5, row.get("taxPeriods",""))
        w(7, row.get("noticeType",""))
        w(8, row.get("noticeDesc",""))
        w(9, row.get("status","OPEN"))
        try: w(10, float(row.get("issueAmt","") or ""))
        except: pass
        try: w(12, float(row.get("amountDue","") or ""))
        except: pass
        w(14, row.get("notes",""))
        w(15, row.get("abgAction",""))

def tracker_rows():
    wb = st.session_state.tracker_wb
    if not wb: return [], set()
    ws = wb["Notices"]
    rows, juris = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        fnum = row[0]
        if not fnum or not any(v for v in row[1:17] if v is not None): continue
        d = row[2]
        if isinstance(d, (date, datetime)): dfmt = d.strftime("%m-%d-%y")
        elif isinstance(d, (int, float)): dfmt = (date(1899,12,30)+timedelta(days=int(d))).strftime("%m-%d-%y")
        else: dfmt = str(d or "")
        j = str(row[4] or "")
        if j: juris.add(j)
        rows.append({"#": int(fnum), "Date": dfmt, "Entity": str(row[3] or ""),
                     "Jurisdiction": j, "Tax Year": str(row[5] or ""),
                     "Type": str(row[7] or ""), "Notice": str(row[8] or ""),
                     "Status": str(row[9] or ""), "$ Issue": row[10]})
    return rows, juris

# ─────────────────────────────────────────────────────────────────────────────
# ── RENDER ────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

# ── Header ────────────────────────────────────────────────────────────────────
next_num = st.session_state.next_file_num
st.markdown(f"""
<div class="abg-header">
  <div class="abg-logo">ABG</div>
  <div>
    <h1>ABG Notice Tracker — Upload Tool</h1>
    <p>Upload IRS / state notices · Extract via Claude · Download updated tracker</p>
  </div>
  <div class="abg-badge">Next #: {next_num}</div>
</div>
<div class="abg-main">
""", unsafe_allow_html=True)

# ── Step row ──────────────────────────────────────────────────────────────────
col1, col2 = st.columns([1.5, 1], gap="small")

# ── Step 1: PDF upload ────────────────────────────────────────────────────────
with col1:
    st.markdown('<div class="step-card">', unsafe_allow_html=True)
    st.markdown('<div class="step-title">① Upload PDF Notices</div>', unsafe_allow_html=True)

    pdfs = st.file_uploader("PDFs", type=["pdf"], accept_multiple_files=True,
                             label_visibility="collapsed", key="pdf_up")
    if pdfs:
        existing = {f["name"] for f in st.session_state.pdf_files}
        for p in pdfs:
            if p.name not in existing:
                n = st.session_state.next_file_num + len(st.session_state.pdf_files)
                st.session_state.pdf_files.append({"name": p.name, "bytes": p.read(), "num": n})
                existing.add(p.name)

    if st.session_state.pdf_files:
        for i, f in enumerate(st.session_state.pdf_files):
            c1, c2, c3 = st.columns([3, 1.2, 0.4])
            c1.markdown(f'<div class="queue-fname">📄 {f["name"]}</div>', unsafe_allow_html=True)
            c2.markdown(f'<div class="queue-badge">→ Notice #{f["num"]}</div>', unsafe_allow_html=True)
            if c3.button("✕", key=f"rm{i}"):
                st.session_state.pdf_files.pop(i)
                for j, ff in enumerate(st.session_state.pdf_files):
                    ff["num"] = st.session_state.next_file_num + j
                st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ── Step 2: Tracker ───────────────────────────────────────────────────────────
with col2:
    st.markdown('<div class="step-card">', unsafe_allow_html=True)
    st.markdown('<div class="step-title">② Tracker (optional)</div>', unsafe_allow_html=True)

    if st.session_state.tracker_label:
        st.markdown(f'<div class="tracker-loaded">✓ {st.session_state.tracker_label}</div>',
                    unsafe_allow_html=True)
    else:
        st.caption("Upload tracker to auto-detect next file number.")

    tf = st.file_uploader(".xlsx", type=["xlsx","xls"],
                           label_visibility="collapsed", key="tracker_up")
    if tf:
        load_tracker(tf.read(), tf.name)
        st.rerun()

    embed_date = st.session_state.get(TRACKER_DATE_KEY, "—")
    st.markdown(f'<p class="tip">Tracker last updated: <strong>{embed_date}</strong></p>',
                unsafe_allow_html=True)

    if st.session_state.get(TRACKER_B64_KEY):
        if st.button("✕ Clear saved tracker", key="clear_t"):
            for k in [TRACKER_B64_KEY, TRACKER_DATE_KEY]:
                st.session_state.pop(k, None)
            st.session_state.tracker_wb    = None
            st.session_state.tracker_label = ""
            st.session_state.next_file_num = 578
            st.rerun()

    st.markdown("---")
    st.markdown('<p class="tip">Update tool with a newer tracker:</p>', unsafe_allow_html=True)
    rf = st.file_uploader("Update Tool", type=["xlsx","xls"],
                           label_visibility="collapsed", key="reembed_up")
    if rf:
        st.info("Use update_tool.py to re-embed this tracker into the HTML file.", icon="ℹ️")

    st.markdown('</div>', unsafe_allow_html=True)

# ── Tracker Status Panel ──────────────────────────────────────────────────────
with st.expander("🗂  View Current Tracker Status", expanded=False):
    rows, juris_set = tracker_rows()
    if not rows:
        st.info("No tracker loaded yet.")
    else:
        fc1, fc2, fc3 = st.columns([3,1,1])
        search  = fc1.text_input("Search", placeholder="Search entity, jurisdiction, notice…",
                                  label_visibility="collapsed", key="ts_s")
        sf      = fc2.selectbox("Status", ["All","OPEN","CLOSED","Blank"],
                                 label_visibility="collapsed", key="ts_st")
        jf_opts = ["All Jurisdictions"] + sorted(juris_set)
        jf      = fc3.selectbox("Jurisdiction", jf_opts,
                                 label_visibility="collapsed", key="ts_j")

        filtered = rows
        if search:
            q = search.lower()
            filtered = [r for r in filtered
                        if q in f"{r['#']} {r['Entity']} {r['Jurisdiction']} {r['Notice']} {r['Type']}".lower()]
        if sf != "All":
            if sf == "Blank": filtered = [r for r in filtered if not r["Status"].strip()]
            else: filtered = [r for r in filtered if sf.upper() in r["Status"].upper()]
        if jf != "All Jurisdictions":
            filtered = [r for r in filtered if r["Jurisdiction"] == jf]

        st.dataframe(pd.DataFrame(filtered), use_container_width=True,
                     height=360, hide_index=True)
        st.caption(f"Showing {len(filtered)} of {len(rows)} notices")

# ── Extract button ────────────────────────────────────────────────────────────
st.markdown("")
can_extract = len(st.session_state.pdf_files) > 0

if st.button("⚡  Extract & Process Notices", type="primary",
             use_container_width=True, disabled=not can_extract):
    st.session_state.logs = []
    st.session_state.extracted_rows = []
    prog = st.progress(0, text="Starting…")
    log_ph = st.empty()
    total = len(st.session_state.pdf_files)

    for i, f in enumerate(st.session_state.pdf_files):
        prog.progress(int(i/total*80), text=f"Processing {f['name']}…")
        log(f"Processing: {f['name']} → Notice #{f['num']}")
        log_ph.markdown(
            '<div class="log-box">' +
            "".join(f'<div class="log-{l}">{m}</div>'
                    for l, m in st.session_state.logs) +
            '</div>', unsafe_allow_html=True)
        try:
            ext = extract(f["bytes"])
            row = {"fileNum": f["num"], "sourceFile": f["name"], "pdfBytes": f["bytes"],
                   "entityName": ext.get("entityName",""),
                   "jurisdiction": ext.get("jurisdiction",""),
                   "taxPeriods": ext.get("taxPeriods",""),
                   "noticeType": ext.get("noticeType","Informational"),
                   "noticeDesc": ext.get("noticeDesc",""),
                   "noticeDate": ext.get("noticeDate",""),
                   "issueAmt": ext.get("issueAmt",""),
                   "amountDue": ext.get("amountDue",""),
                   "abgAction": ext.get("abgAction",""),
                   "notes": ext.get("notes",""),
                   "status": "OPEN"}
            st.session_state.extracted_rows.append(row)
            log(f"✓ {f['name']} → {row['entityName']} | {row['jurisdiction']} | {row['taxPeriods']}", "ok")
        except Exception as e:
            log(f"✗ Failed: {f['name']} — {e}", "err")
            st.session_state.extracted_rows.append({
                "fileNum": f["num"], "sourceFile": f["name"], "pdfBytes": f["bytes"],
                "entityName":"","jurisdiction":"","taxPeriods":"",
                "noticeType":"Informational","noticeDesc":"",
                "noticeDate":"","issueAmt":"","amountDue":"",
                "abgAction":"","notes": f["name"],"status":"OPEN"})

    prog.progress(100, text="Done!")
    log(f"Extracted {len(st.session_state.extracted_rows)} notice(s).", "ok")
    st.rerun()

# ── Log box ───────────────────────────────────────────────────────────────────
if st.session_state.logs:
    st.markdown(
        '<div class="log-box">' +
        "".join(f'<div class="log-{l}">{m}</div>'
                for l, m in st.session_state.logs[-30:]) +
        '</div>', unsafe_allow_html=True)

# ── Results ───────────────────────────────────────────────────────────────────
if st.session_state.extracted_rows:
    n = len(st.session_state.extracted_rows)
    st.markdown(f"""
    <div class="results-header" style="margin-top:24px;">
      <h2>Extracted Notices</h2>
      <span>{n} row{"s" if n!=1 else ""} — review and edit before downloading</span>
    </div>
    """, unsafe_allow_html=True)

    display_cols = ["fileNum","noticeDate","entityName","jurisdiction","taxPeriods",
                    "noticeType","noticeDesc","status","issueAmt","amountDue","abgAction","notes"]
    col_labels   = ["File #","Date","Entity","Jurisdiction","Tax Year",
                    "Type","Notice","Status","$ Issue","Amount Due","ABG Action","Notes"]

    df = pd.DataFrame([{l: r.get(k,"") for l,k in zip(col_labels,display_cols)}
                       for r in st.session_state.extracted_rows])
    edited = st.data_editor(df, use_container_width=True, num_rows="dynamic", hide_index=True)

    for i, row in enumerate(st.session_state.extracted_rows):
        if i < len(edited):
            for l, k in zip(col_labels, display_cols):
                row[k] = edited.iloc[i][l]

    # Action buttons row — matches HTML: + Add Row | ↻ Re-extract | ✕ Clear All
    ba1, ba2, ba3 = st.columns(3)
    with ba1:
        if st.button("＋  Add Row", use_container_width=True):
            new_num = (max(r["fileNum"] for r in st.session_state.extracted_rows) + 1
                       if st.session_state.extracted_rows
                       else st.session_state.next_file_num)
            st.session_state.extracted_rows.append({
                "fileNum": new_num, "entityName": "", "jurisdiction": "",
                "taxPeriods": "", "noticeType": "Informational", "noticeDesc": "",
                "noticeDate": "", "status": "OPEN", "issueAmt": "", "amountDue": "",
                "abgAction": "", "notes": "", "sourceFile": "manual", "pdfBytes": None,
            })
            st.rerun()
    with ba2:
        if st.button("↻  Re-extract", use_container_width=True):
            # Re-run extraction on any PDFs that were uploaded (skip manual rows)
            st.session_state.extracted_rows = [
                r for r in st.session_state.extracted_rows if r.get("sourceFile") == "manual"
            ]
            st.session_state.logs = []
            st.session_state["_reextract"] = True
            st.rerun()
    with ba3:
        if st.button("✕  Clear All", use_container_width=True, type="secondary"):
            if st.session_state.get("_confirm_clear"):
                st.session_state.pdf_files = []
                st.session_state.extracted_rows = []
                st.session_state.logs = []
                st.session_state.pop("_confirm_clear", None)
                st.rerun()
            else:
                st.session_state["_confirm_clear"] = True
                st.rerun()

    if st.session_state.get("_confirm_clear"):
        st.warning("Click ✕ Clear All again to confirm.")

    # Handle re-extract trigger
    if st.session_state.get("_reextract"):
        st.session_state.pop("_reextract", None)
        prog = st.progress(0, text="Re-extracting…")
        log_ph = st.empty()
        pdfs_to_process = [f for f in st.session_state.pdf_files]
        total = len(pdfs_to_process)
        if total == 0:
            log("No PDFs to re-extract.", "warn")
        for i, f in enumerate(pdfs_to_process):
            prog.progress(int(i/total*80) if total else 80, text=f"Processing {f['name']}…")
            log(f"Re-extracting: {f['name']} → Notice #{f['num']}")
            log_ph.markdown(
                '<div class="log-box">' +
                "".join(f'<div class="log-{l}">{m}</div>' for l, m in st.session_state.logs) +
                '</div>', unsafe_allow_html=True)
            try:
                ext = extract(f["bytes"])
                row = {"fileNum": f["num"], "sourceFile": f["name"], "pdfBytes": f["bytes"],
                       "entityName": ext.get("entityName",""),
                       "jurisdiction": ext.get("jurisdiction",""),
                       "taxPeriods": ext.get("taxPeriods",""),
                       "noticeType": ext.get("noticeType","Informational"),
                       "noticeDesc": ext.get("noticeDesc",""),
                       "noticeDate": ext.get("noticeDate",""),
                       "issueAmt": ext.get("issueAmt",""),
                       "amountDue": ext.get("amountDue",""),
                       "abgAction": ext.get("abgAction",""),
                       "notes": ext.get("notes",""),
                       "status": "OPEN"}
                st.session_state.extracted_rows.append(row)
                log(f"✓ {f['name']} → {row['entityName']} | {row['jurisdiction']}", "ok")
            except Exception as e:
                log(f"✗ Failed: {f['name']} — {e}", "err")
        prog.progress(100, text="Done!")
        st.rerun()

    # Summary strip
    total_due = 0
    for r in st.session_state.extracted_rows:
        try: total_due += float(r.get("amountDue") or r.get("issueAmt") or 0)
        except: pass
    juris_c = len({r.get("jurisdiction") for r in st.session_state.extracted_rows if r.get("jurisdiction")})

    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Notices", n)
    m2.metric("Total Due", f"${total_due:,.0f}" if total_due else "$0")
    m3.metric("Jurisdictions", juris_c)
    m4.metric("Starting File #", f"#{st.session_state.next_file_num}")

    # Download section
    st.markdown(f"""
    <div class="download-section" style="margin-top:20px;">
      <h3>✓ Ready to Download</h3>
      <p>All extracted rows will be written to the tracker starting at
         <strong>File #{st.session_state.next_file_num}</strong>.
         Review the table above before downloading.</p>
    </div>
    """, unsafe_allow_html=True)

    dc1, dc2, dc3 = st.columns(3)

    with dc1:
        if st.button("⬇  Download Updated Tracker", type="primary", use_container_width=True):
            wb = st.session_state.tracker_wb
            if wb is None:
                wb = openpyxl.Workbook(); ws = wb.active
                ws.title = "Notices"; ws.append(COL_HEADERS)
            write_rows(wb, st.session_state.extracted_rows)
            xb = wb_bytes(wb)
            today = date.today().strftime("%Y-%m-%d")
            st.session_state[TRACKER_B64_KEY]  = base64.b64encode(xb).decode()
            st.session_state[TRACKER_DATE_KEY] = date.today().strftime("%m-%d-%y")
            st.session_state.tracker_wb        = load_wb(xb)
            st.session_state.next_file_num     = get_next_num(st.session_state.tracker_wb["Notices"])
            st.session_state.tracker_label     = f"Updated {date.today().strftime('%m-%d-%y')}"
            st.download_button("⬇  Save Tracker", data=xb,
                               file_name=f"ABG_Notice_Tracker_Updated_{today}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
            log(f"Tracker saved — {today}", "ok")

    with dc2:
        if st.button("⬇  Save Renamed PDFs", use_container_width=True):
            for row in st.session_state.extracted_rows:
                if row.get("pdfBytes"):
                    st.download_button(
                        f"⬇  Notice #{row['fileNum']}.pdf",
                        data=row["pdfBytes"],
                        file_name=f"Notice #{row['fileNum']}.pdf",
                        mime="application/pdf",
                        key=f"dlp_{row['fileNum']}",
                        use_container_width=True)

    with dc3:
        pass  # Clear All is now in the action row above

st.markdown('</div>', unsafe_allow_html=True)
